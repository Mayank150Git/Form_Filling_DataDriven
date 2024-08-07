import openpyxl


class FormDataDriven:

    @staticmethod
    def excel_data1():

        book = openpyxl.load_workbook("C:\\Users\\HP\\Desktop\\Form_Filling_DummyData.xlsx")
        sheet = book.active
        data_list = []

        for i in range(2, sheet.max_row+1):
            Dict = {}
            for j in range(1, sheet.max_column+1):
                Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            data_list.append(Dict)
        return data_list



